import random
import statistics
from openpyxl import Workbook
from openpyxl import load_workbook


class Customer:

    def __init__(self, arrivalTime, num):
        self.arrivalTime = arrivalTime
        self.serviceTime = 0
        self.waitingTime = 0
        self.num = num


class Server:

    def __init__(self, serviceTimes, serviceDists):
        self.serviceTimes = serviceTimes
        self.serviceDists = serviceDists
        self.idleTime = 0
        self.customerInService = 0
        self.customersServed = []

    def GenerateServiceTime(self):
        rNum = random.randint(1, 100)
        for i in range(0, len(self.serviceTimes)):
            if (self.customerInService.serviceTime == 0 and rNum <= self.serviceDists[i] * 100):
                self.customerInService.serviceTime = self.serviceTimes[i]


class System:

    def __init__(self, minArrivalTime, maxArrivalTime, numCustomers, numServers, numQs):
        self.minArrivalTime = minArrivalTime
        self.maxArrivalTime = maxArrivalTime
        self.numCustomers = numCustomers
        self.numServers = numServers
        self.numQs = numQs

    def GenerateCustomers(self, CustomersList):
        arvlT = 0
        for i in range(0, self.numCustomers):
            CustomersList.append(Customer(arvlT, i + 1))
            arvlT += random.randint(self.minArrivalTime, self.maxArrivalTime)


class simStats:

    def __init__(self, cList, sList, systemNum):
        self.cList = cList
        self.sList = sList
        self.systemNum = systemNum

    def avgWaitingTime(self):
        cWT = []
        for c in self.cList:
            cWT.append(c.waitingTime)
        return statistics.mean(cWT)

    def pWaiting(self):
        cWait = 0
        for c in self.cList:
            if (c.waitingTime > 0):
                cWait += 1
        return cWait / len(self.cList)

    def pServerIdle(self):
        sIdle = 0
        for s in self.sList:
            sIdle += s.idleTime
        lastC = self.cList[-1]
        return sIdle / (lastC.arrivalTime + lastC.waitingTime + lastC.serviceTime) / len(self.sList)

    def avgServiceTime(self):
        cST = []
        for c in self.cList:
            cST.append(c.serviceTime)
        return statistics.mean(cST)

    def avgInterArrivalTime(self):
        cIAT = []
        for i in range(1, len(self.cList)):
            cIAT.append(self.cList[i].arrivalTime -
                        self.cList[i - 1].arrivalTime)
        return statistics.mean(cIAT)

    def avgWaitingTime2(self):
        cWT = []
        for c in self.cList:
            if(c.waitingTime > 0):
                cWT.append(c.waitingTime)
        if(len(cWT) > 0):
            return statistics.mean(cWT)
        else:
            return 0

    def avgTimeSystem(self):
        return self.avgServiceTime() + self.avgWaitingTime()

    def stdServiceTime(self):
        cST = []
        for c in self.cList:
            cST.append(c.serviceTime)
        return statistics.stdev(cST)

    def stdWaitingTime(self):
        cWT = []
        for c in self.cList:
            cWT.append(c.waitingTime)
        return statistics.stdev(cWT)

    def stdInterArrivalTime(self):
        cIAT = []
        for i in range(1, len(self.cList)):
            cIAT.append(self.cList[i].arrivalTime -
                        self.cList[i - 1].arrivalTime)
        return statistics.stdev(cIAT)

    def taNumInSystem(self):
        return statistics.mean(self.systemNum)

    def singleCTrialResults(self):
        singleCTrials = []
        for s in self.sList:
            singleCTrials.append([])
            for i in range(0, len(s.customersServed)):
                cs = s.customersServed[i]
                if (cs.num == 1):
                    singleCTrials[-1].append([cs.num, -1, cs.arrivalTime, cs.serviceTime,
                                              cs.waitingTime + cs.arrivalTime, cs.waitingTime,
                                              cs.waitingTime + cs.arrivalTime + cs.serviceTime,
                                              cs.waitingTime + cs.serviceTime, -1, self.sList.index(s) + 1])
                elif(i == 0):
                    singleCTrials[-1].append([cs.num, cs.arrivalTime - self.cList[self.cList.index(cs) - 1].arrivalTime,
                                              cs.arrivalTime, cs.serviceTime, cs.waitingTime + cs.arrivalTime,
                                              cs.waitingTime, cs.waitingTime + cs.arrivalTime + cs.serviceTime,
                                              cs.waitingTime + cs.serviceTime, cs.waitingTime + cs.arrivalTime,
                                              self.sList.index(s) + 1])
                else:
                    csOld = s.customersServed[i - 1]
                    singleCTrials[-1].append([cs.num, cs.arrivalTime - self.cList[self.cList.index(cs) - 1].arrivalTime,
                                              cs.arrivalTime, cs.serviceTime, cs.waitingTime + cs.arrivalTime,
                                              cs.waitingTime, cs.waitingTime + cs.arrivalTime + cs.serviceTime,
                                              cs.waitingTime + cs.serviceTime, cs.waitingTime + cs.arrivalTime -
                                              csOld.waitingTime - csOld.arrivalTime - csOld.serviceTime,
                                              self.sList.index(s) + 1])
        return singleCTrials


def avgMultipleTrialStats(multipleTrialStats):
    tempStats = []
    for stat in multipleTrialStats[0]:
        tempStats.append([])
    for statList in multipleTrialStats:
        for j in range(0, len(statList)):
            tempStats[j].append(statList[j])
    for i in range(0, len(tempStats)):
        tempStats[i] = statistics.mean(tempStats[i])
    return tempStats


def simOutput(wsx, caseData):
    wsx['A1'] = 'Customer Number'
    wsx['B1'] = 'InterArrival Time'
    wsx['C1'] = 'Arrival Time'
    wsx['D1'] = 'Service Time'
    wsx['E1'] = 'Time Service Begins'
    wsx['F1'] = 'Waiting Time in Queue'
    wsx['G1'] = 'Time Service Ends'
    wsx['H1'] = 'Time Customer Spends in System'
    wsx['I1'] = 'Idle Time of Server'
    wsx['J1'] = 'Server Number'
    custCount = 2
    for s in caseData:
        for cust in s:
            for i in range(1, len(cust) + 1):
                if cust[i - 1] >= 0:
                    wsx.cell(row=custCount, column=i).value = cust[i - 1]
            custCount += 1


def runSim(case, minArrivalTime, maxArrivalTime, numCustomers, numServers, serviceTimesList, serviceDistributionList):
    CustomersList = []
    currentTime = 0
    custNum = 0
    numInSystem = []
    if (case == 1):
        Queue = []
        Servers = []
        Sys = System(minArrivalTime, maxArrivalTime, numCustomers, 1, 1)
        Servers.append(Server(serviceTimesList, serviceDistributionList))
    elif (case == 2):
        Queue = []
        Servers = []
        Sys = System(minArrivalTime, maxArrivalTime,
                     numCustomers, numServers, 1)
        for i in range(0, Sys.numServers):
            Servers.append(Server(serviceTimesList, serviceDistributionList))
    elif (case == 3 or case == 4):
        Queues = []
        Servers = []
        Sys = System(minArrivalTime, maxArrivalTime,
                     numCustomers, numServers, numServers)
        for i in range(0, Sys.numServers):
            Servers.append(Server(serviceTimesList, serviceDistributionList))
            Queues.append([])
    Sys.GenerateCustomers(CustomersList)

    if (case == 1 or case == 2):
        while(CustomersList[-1].serviceTime == 0):
            numInSystem.append(0)
            if (CustomersList[custNum].arrivalTime == currentTime):
                Queue.append(CustomersList[custNum])
                custNum = min(custNum + 1, Sys.numCustomers - 1)
            for s in Servers:
                if(s.customerInService != 0 and s.customerInService.arrivalTime + s.customerInService.serviceTime +
                   s.customerInService.waitingTime == currentTime):
                    s.customerInService = 0
            for s in Servers:
                if (s.customerInService == 0 and len(Queue) > 0):
                    s.customerInService = Queue.pop(0)
                    s.GenerateServiceTime()
                    s.customersServed.append(s.customerInService)
            for Person in Queue:
                Person.waitingTime += 1
            for s in Servers:
                if (s.customerInService == 0):
                    s.idleTime += 1
                else:
                    numInSystem[-1] += 1
            numInSystem[-1] += len(Queue)
            currentTime += 1
    elif (case == 3 or case == 4):
        while(CustomersList[-1].serviceTime == 0):
            numInSystem.append(0)
            if (CustomersList[custNum].arrivalTime == currentTime):
                Queues[random.randint(0, len(Queues) - 1)
                       ].append(CustomersList[custNum])
                custNum = min(custNum + 1, Sys.numCustomers - 1)
            for s in Servers:
                if(s.customerInService != 0 and s.customerInService.arrivalTime +
                   s.customerInService.serviceTime + s.customerInService.waitingTime == currentTime):
                    s.customerInService = 0
            for s in Servers:
                if (case == 3 and s.customerInService == 0 and len(Queues[Servers.index(s)]) > 0):
                    s.customerInService = Queues[Servers.index(s)].pop(0)
                    s.GenerateServiceTime()
                    s.customersServed.append(s.customerInService)
                if (case == 4 and s.customerInService == 0):
                    if (len(Queues[Servers.index(s)]) > 0):
                        s.customerInService = Queues[Servers.index(s)].pop(0)
                        s.GenerateServiceTime()
                        s.customersServed.append(s.customerInService)
                    else:
                        NonEmptyQueuesIndices = []
                        for i in range(0, len(Queues)):
                            if (len(Queues[i]) > 0):
                                NonEmptyQueuesIndices.append(i)
                        if (len(NonEmptyQueuesIndices) > 0):
                            s.customerInService = Queues[
                                random.choice(NonEmptyQueuesIndices)].pop(0)
                            s.GenerateServiceTime()
                            s.customersServed.append(s.customerInService)
            for Queue in Queues:
                numInSystem[-1] += len(Queue)
                for Person in Queue:
                    Person.waitingTime += 1
            for s in Servers:
                if (s.customerInService == 0):
                    s.idleTime += 1
                else:
                    numInSystem[-1] += 1
            currentTime += 1

    sS = simStats(CustomersList, Servers, numInSystem)

    if (case == 1):
        simOutput(ws1, sS.singleCTrialResults())
    elif (case == 2):
        simOutput(ws2, sS.singleCTrialResults())
    elif (case == 3):
        simOutput(ws3, sS.singleCTrialResults())
    elif (case == 4):
        simOutput(ws4, sS.singleCTrialResults())

    multipleTrialStatsLocal.append([sS.avgWaitingTime(), sS.pWaiting(), sS.pServerIdle(), sS.avgServiceTime(),
                                    1 - sS.pServerIdle(), sS.avgInterArrivalTime(), sS.avgWaitingTime2(),
                                    sS.avgTimeSystem(), sS.stdInterArrivalTime(), sS.taNumInSystem(),
                                    sS.stdWaitingTime(), sS.stdServiceTime(), case])
    allTrialData.append(multipleTrialStatsLocal[-1])


random.seed()
wbIn = load_workbook(filename='SimIn.xlsx', data_only=True)
ws = wbIn['Sheet1']

inNumCustomers = ws['B1'].value
inNumServers = ws['B2'].value
inNumTrials = ws['B3'].value
inMinArrivalTime = ws['D1'].value
inMaxArrivalTime = ws['D2'].value
inServiceTimes = []
for i in range(1, len(ws.columns[4])):
    inServiceTimes.append(ws.columns[4][i].value)
inServiceDistribution = []
for i in range(1, len(ws.columns[6])):
    inServiceDistribution.append(ws.columns[6][i].value)
multipleTrialStatsCase = []
allTrialData = []

wbOut = Workbook()
ws0 = wbOut.active
ws0.title = 'Multiple Trial Stats'
ws1 = wbOut.create_sheet()
ws1.title = 'Case 1'
ws2 = wbOut.create_sheet()
ws2.title = 'Case 2'
ws3 = wbOut.create_sheet()
ws3.title = 'Case 3'
ws4 = wbOut.create_sheet()
ws4.title = 'Case 4'

for caseNum in range(1, 5):
    multipleTrialStatsLocal = []
    for i in (0, inNumTrials):
        runSim(caseNum, inMinArrivalTime, inMaxArrivalTime, inNumCustomers,
               inNumServers, inServiceTimes, inServiceDistribution)
    multipleTrialStatsCase.append(
        avgMultipleTrialStats(multipleTrialStatsLocal))
ws0['A1'] = 'Trial'
ws0['B1'] = 'Average Customer Waiting Time'
ws0['C1'] = 'Probability of Waiting'
ws0['D1'] = 'Probability of Server Being Idle'
ws0['E1'] = 'Average Service Time'
ws0['F1'] = 'Average Server Utilization'
ws0['G1'] = 'Average Time Between Arrivals'
ws0['H1'] = 'Average Waiting Time of Those Who Wait'
ws0['I1'] = 'Average Time Customer Spends in the System'
ws0['J1'] = 'Standard Deviation of Inter-Arrival Time'
ws0['K1'] = 'Time-Average Number in System'
ws0['L1'] = 'Standard Deviation of Waiting Time'
ws0['M1'] = 'Standard Deviation of Service Time'
ws0['N1'] = 'Case Number'

for i in range(2, len(allTrialData) + 2):
    ws0.cell(row=i, column=1).value = i - 1
    for j in range(2, len(allTrialData[0]) + 2):
        ws0.cell(row=i, column=j).value = allTrialData[i - 2][j - 2]

rowOffset = len(allTrialData)
for i in range(rowOffset + 2, len(multipleTrialStatsCase) + rowOffset + 2):
    ws0.cell(row=i, column=1).value = 'Average'
    for j in range(2, len(multipleTrialStatsCase[0]) + 2):
        ws0.cell(row=i, column=j).value = multipleTrialStatsCase[
            i - 2 - rowOffset][j - 2]

wbOut.save('SimOut.xlsx')
